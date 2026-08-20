import math
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DETAILED_COLUMNS = [
    "sheet", "source_row", "medication_column", "patient", "original", "drug", "dose_mg", "frequency",
    "daily_dose_mg", "method", "target_drug", "equivalent_dose_mg", "warning",
    "match_type", "match_score", "needs_review", "method_warning",
]
AUDIT_COLUMNS = ["sheet", "source_row", "medication_column", "patient", "original", "parsed", "match_type", "match_score", "status", "unavailable_methods"]
ERROR_COLUMNS = ["sheet", "source_row", "medication_column", "patient", "original", "error"]
METHOD_INFO = [
    {"method": "CMD", "basis": "Classical mean dose method", "reference": "Leucht et al. 2015; PMID 25841041", "limitation": "급성 조현병 경구약 임상시험 평균용량 기반이며 개인별 처방 권고가 아님"},
    {"method": "MED", "basis": "Minimum effective dose method", "reference": "Leucht et al. 2014; PMID 24493852", "limitation": "초발성·치료저항성 환자에 일반화할 수 없음"},
    {"method": "ED95", "basis": "95% effective dose method", "reference": "Leucht et al. 2020; PMID 31838873", "limitation": "만성 조현병 급성 악화 집단의 평균 효과 기반; haloperidol 값은 단일 연구 기반 제한적 추정치"},
    {"method": "DDD", "basis": "WHO Defined Daily Dose", "reference": "WHO ATC/DDD methodology", "limitation": "약물사용 연구용 기술 단위이며 권장·처방 용량이 아님"},
    {"method": "CPZ_FGA", "basis": "Historical chlorpromazine equivalents", "reference": "Davis 1974; PMID 4156792", "limitation": "1세대 항정신병약물의 역사적 비교값"},
]


def _safe_excel_value(value):
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _safe_frame(rows, columns):
    frame = pd.DataFrame(rows, columns=columns)
    return frame.map(_safe_excel_value)


def _format_worksheet(worksheet):
    header_fill = PatternFill("solid", fgColor="DCEBFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="12233F")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 30
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    wrap_headers = {"original", "warning", "method_warning", "error", "limitation", "reference"}
    for column_index, cells in enumerate(worksheet.iter_cols(), start=1):
        header = str(cells[0].value or "")
        max_length = max((len(str(cell.value)) for cell in cells if cell.value is not None), default=0)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 11), 55)
        for cell in cells[1:]:
            cell.alignment = Alignment(vertical="top", wrap_text=header in wrap_headers)

    headers = {cell.column: str(cell.value or "") for cell in worksheet[1]}
    for row_index in range(2, worksheet.max_row + 1):
        required_lines = 1
        for cell in worksheet[row_index]:
            if headers.get(cell.column) not in wrap_headers or cell.value is None:
                continue
            width = worksheet.column_dimensions[get_column_letter(cell.column)].width or 11
            visual_length = sum(2 if ord(character) > 127 else 1 for character in str(cell.value))
            required_lines = max(required_lines, math.ceil(visual_length / max(int(width), 1)))
        worksheet.row_dimensions[row_index].height = min(max(18, required_lines * 16), 96)


def export_results(detailed_rows, audit_rows, error_rows, directory):
    output_file = Path(directory) / "result.xlsx"
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        _safe_frame(detailed_rows, DETAILED_COLUMNS).to_excel(
            writer, sheet_name="Detailed", index=False
        )
        _safe_frame(audit_rows, AUDIT_COLUMNS).to_excel(
            writer, sheet_name="AuditTrail", index=False
        )
        _safe_frame(error_rows, ERROR_COLUMNS).to_excel(
            writer, sheet_name="Errors", index=False
        )
        pd.DataFrame(METHOD_INFO).to_excel(writer, sheet_name="MethodInfo", index=False)
        for worksheet in writer.book.worksheets:
            _format_worksheet(worksheet)
    return output_file
