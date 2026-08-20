import io
import os
import re
import tempfile
from pathlib import Path

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.utils import secure_filename

from services.anonymizer import anonymize_dataframe
from services.column_detector import detect_columns
from services.converter import available_methods, convert_drug, method_warning, normalize_target
from services.exporter import export_results
from services.file_reader import read_file
from services.medication_splitter import split_medications
from services.parser import DOSE_RE, parse_medication
from services.validator import validate_file


BASE_DIR = Path(__file__).resolve().parent
app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024
METHODS = available_methods()


def _cell_text(value):
    text = "" if value is None else str(value).strip()
    return "" if text.casefold() in {"", "nan", "none", "null"} else text


def _compose_structured_medication(row, drug_text, dose_col, unit_col, frequency_col):
    dose = _cell_text(row.get(dose_col)) if dose_col is not None else ""
    unit = _cell_text(row.get(unit_col)) if unit_col is not None else ""
    frequency = _cell_text(row.get(frequency_col)) if frequency_col is not None else ""
    dose_header = str(dose_col or "").casefold()
    if not unit:
        unit_match = re.search(r"(?:^|[_\s(])(mcg|ug|μg|㎍|mg|㎎|g)(?:$|[_\s)])", dose_header)
        unit = unit_match.group(1) if unit_match else ""
    is_daily_dose = any(word in dose_header for word in ("daily", "일일", "1일"))
    if is_daily_dose:
        frequency = "QD"
    else:
        normalized_frequency = frequency.replace(".0", "")
        if normalized_frequency == "0.5":
            frequency = "QOD"
        elif normalized_frequency.isdigit():
            count = int(normalized_frequency)
            frequency = {1: "QD", 2: "BID", 3: "TID", 4: "QID"}.get(count, f"{count} times a day")
    return " ".join(part for part in (drug_text, f"{dose}{unit}" if dose else "", frequency) if part)


@app.get("/")
def home():
    return render_template("index.html", methods=METHODS)


@app.get("/health")
def health():
    return {"status": "ok"}


@app.get("/sample")
def sample_file():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "입력예시"
    sheet.append(["patient_id", "drug", "dose", "unit", "frequency"])
    sheet.append(["P001", "Risperdal", 2, "mg", "BID"])
    sheet.append(["P002", "Abilify", 15, "mg", "QD"])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1769E0")
        cell.alignment = Alignment(horizontal="center")
    for column, width in {"A": 16, "B": 20, "C": 12, "D": 10, "E": 16}.items():
        sheet.column_dimensions[column].width = width
    sample = io.BytesIO()
    workbook.save(sample)
    sample.seek(0)
    return send_file(
        sample,
        as_attachment=True,
        download_name="AP_equivalence_sample.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/api/parse")
def parse_text():
    payload = request.get_json(silent=True) or {}
    text = str(payload.get("text", "")).strip()
    if not text:
        return jsonify({"error": "약물과 용량을 입력해 주세요."}), 400
    items = []
    for medication in split_medications(text):
        try:
            parsed = parse_medication(medication)
            conversions = []
            for method in METHODS:
                try:
                    target = normalize_target(method)
                    equivalent = convert_drug(parsed["drug"], parsed["daily_dose_mg"], method, target)
                    conversions.append({"method": method, "target": target, "value": round(equivalent, 4)})
                except (ValueError, LookupError):
                    conversions.append({"method": method, "target": "", "value": None})
            items.append({"ok": True, **parsed, "conversions": conversions})
        except ValueError as exc:
            items.append({"ok": False, "original": medication, "error": str(exc), "needs_review": True})
    return jsonify({"items": items})


@app.post("/upload")
def upload():
    uploaded_file = request.files.get("file")
    method = request.form.get("method", "ALL").strip().upper() or "ALL"

    try:
        validate_file(uploaded_file)
        selected_methods = METHODS if method == "ALL" else [method]
        for selected_method in selected_methods:
            normalize_target(selected_method)
    except ValueError as exc:
        return render_template("error.html", message=str(exc)), 400

    original_suffix = Path(uploaded_file.filename).suffix.lower()
    filename = secure_filename(uploaded_file.filename) or f"upload{original_suffix}"
    suffix = original_suffix or Path(filename).suffix.lower()
    detailed_rows, audit_rows, error_rows = [], [], []

    try:
        with tempfile.TemporaryDirectory() as temp_dir:
            upload_path = Path(temp_dir) / f"upload{suffix}"
            uploaded_file.save(upload_path)
            sheets = read_file(str(upload_path))

            patient_mapping = {}
            for sheet_name, dataframe in sheets.items():
                detected = detect_columns(dataframe)
                patient_col = detected["patient_column"]
                medication_col = detected["medication_column"]
                dose_col = detected["dose_column"]
                unit_col = detected["unit_column"]
                frequency_col = detected["frequency_column"]

                if medication_col is None:
                    error_rows.append({"sheet": sheet_name, "source_row": "", "medication_column": "", "patient": "", "original": "", "error": "약물 열을 찾지 못했습니다."})
                    continue

                anonymized, patient_mapping = anonymize_dataframe(
                    dataframe,
                    columns=[patient_col] if patient_col is not None else [],
                    mapping=patient_mapping,
                )
                header_row = int(dataframe.attrs.get("header_row", 0))
                for row_index, row in anonymized.iterrows():
                    patient_id = row.get(patient_col, "") if patient_col is not None else ""
                    source_row = int(row_index) + header_row + 2
                    raw_medication = _cell_text(row.get(medication_col))
                    medications = split_medications(raw_medication)
                    if medications and dose_col is not None and not any(DOSE_RE.search(item) for item in medications):
                        if len(medications) > 1:
                            error_rows.append({"sheet": sheet_name, "source_row": source_row, "medication_column": str(medication_col), "patient": patient_id, "original": raw_medication, "error": "여러 약물이 한 셀에 있고 용량이 별도 열에 있어 자동 결합할 수 없습니다."})
                            continue
                        medications = [_compose_structured_medication(row, medications[0], dose_col, unit_col, frequency_col)]
                    for medication in medications:
                        try:
                            parsed = parse_medication(medication)
                            converted_count = 0
                            unavailable_methods = []
                            for selected_method in selected_methods:
                                target = normalize_target(selected_method)
                                try:
                                    equivalent = convert_drug(parsed["drug"], parsed["daily_dose_mg"], selected_method, target)
                                except LookupError:
                                    unavailable_methods.append(selected_method)
                                    continue
                                converted_count += 1
                                detailed_rows.append({
                                    "sheet": sheet_name,
                                    "source_row": source_row,
                                    "medication_column": str(medication_col),
                                    "patient": patient_id,
                                    "original": medication,
                                    "drug": parsed["drug"],
                                    "dose_mg": parsed["dose_mg"],
                                    "frequency": parsed["frequency"],
                                    "daily_dose_mg": parsed["daily_dose_mg"],
                                    "method": selected_method,
                                    "target_drug": target,
                                    "equivalent_dose_mg": round(equivalent, 4),
                                    "warning": parsed["warning"],
                                    "match_type": parsed["match_type"],
                                    "match_score": round(parsed["match_score"], 1),
                                    "needs_review": parsed["needs_review"],
                                    "method_warning": method_warning(selected_method, parsed["drug"], target),
                                })
                            if converted_count == 0:
                                raise LookupError(f"{parsed['drug']}에 사용할 수 있는 환산값이 없습니다.")
                            audit_rows.append({"sheet": sheet_name, "source_row": source_row, "medication_column": str(medication_col), "patient": patient_id, "original": medication, "parsed": parsed["drug"], "match_type": parsed["match_type"], "match_score": round(parsed["match_score"], 1), "status": "review" if parsed["needs_review"] else "converted", "unavailable_methods": ", ".join(unavailable_methods)})
                        except (ValueError, LookupError) as exc:
                            error_rows.append({"sheet": sheet_name, "source_row": source_row, "medication_column": str(medication_col), "patient": patient_id, "original": medication, "error": str(exc)})

            output_file = export_results(detailed_rows, audit_rows, error_rows, directory=temp_dir)
            result_bytes = io.BytesIO(Path(output_file).read_bytes())
            result_bytes.seek(0)
            return send_file(
                result_bytes,
                as_attachment=True,
                download_name="AP_equivalence_result.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    except (ValueError, OSError) as exc:
        return render_template("error.html", message=str(exc)), 400


@app.errorhandler(413)
def file_too_large(_error):
    return render_template("error.html", message="파일 크기는 최대 100MB입니다."), 413


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")), debug=False)
