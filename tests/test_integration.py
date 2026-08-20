import io
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app import app
from services.parser import parse_medication


def _upload(client, content, filename, method="CMD"):
    return client.post(
        "/upload",
        data={"method": method, "file": (io.BytesIO(content), filename)},
        content_type="multipart/form-data",
    )


def test_korean_cp949_csv_is_detected_and_converted():
    client = app.test_client()
    csv_bytes = "환자번호,처방내역\nA-001,리스페달 2mg BID\nA-002,아빌리파이 15mg QD\n".encode("cp949")
    response = _upload(client, csv_bytes, "처방.csv")
    assert response.status_code == 200
    detailed = pd.read_excel(io.BytesIO(response.data), sheet_name="Detailed")
    assert detailed["drug"].tolist() == ["risperidone", "aripiprazole"]
    assert detailed["patient"].tolist() == ["PATIENT_00001", "PATIENT_00002"]
    assert detailed["daily_dose_mg"].tolist() == [4, 15]


def test_offset_header_multisheet_xlsx_and_review_rows():
    workbook = io.BytesIO()
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        pd.DataFrame({"환자번호": ["P01", "P02"], "약물정보": ["Risperdal 2㎎ BID", "lithium 600mg"]}).to_excel(
            writer, sheet_name="처방", index=False, startrow=2
        )
        pd.DataFrame({"설명": ["처방 데이터가 아닌 시트"]}).to_excel(writer, sheet_name="안내", index=False)
    response = _upload(app.test_client(), workbook.getvalue(), "실제처방.xlsx")
    assert response.status_code == 200
    detailed = pd.read_excel(io.BytesIO(response.data), sheet_name="Detailed")
    errors = pd.read_excel(io.BytesIO(response.data), sheet_name="Errors")
    assert detailed.loc[0, "drug"] == "risperidone"
    assert detailed.loc[0, "daily_dose_mg"] == 4
    assert errors["original"].fillna("").str.casefold().str.contains("lithium").any()


def test_parser_api_marks_unknown_drugs_for_review():
    response = app.test_client().post("/api/parse", json={"text": "Risperdal 2mg BID, lithium 600mg"})
    assert response.status_code == 200
    items = response.get_json()["items"]
    assert items[0]["drug"] == "risperidone" and items[0]["needs_review"] is False
    assert items[1]["ok"] is False and items[1]["needs_review"] is True


def test_ambiguous_schedules_are_not_presented_as_certain():
    client = app.test_client()
    response = client.post("/api/parse", json={"text": "Seroquel 25mg QHS, Risperdal 2mg, Abilify 10mg PRN"})
    items = response.get_json()["items"]
    assert items[0]["frequency"] == "QHS" and items[0]["needs_review"] is False
    assert items[1]["frequency"] == "ASSUMED_QD" and items[1]["needs_review"] is True
    assert items[2]["ok"] is False and "PRN" in items[2]["error"]


def test_multiple_doses_conflicting_frequencies_and_injections_are_rejected():
    text = "risperidone 1mg AM 2mg PM, risperidone 2mg BID TID, risperidone 25mg IM monthly"
    items = app.test_client().post("/api/parse", json={"text": text}).get_json()["items"]
    assert all(item["ok"] is False for item in items)
    assert "용량이 여러 개" in items[0]["error"]
    assert "서로 다른 복용 빈도" in items[1]["error"]
    assert "주사제" in items[2]["error"]


def test_leading_decimal_and_korean_milligram_symbol_are_supported():
    items = app.test_client().post("/api/parse", json={"text": "risperidone .5mg QD, 리스페달 0.5㎎ BID"}).get_json()["items"]
    assert items[0]["dose_mg"] == 0.5 and items[0]["daily_dose_mg"] == 0.5
    assert items[1]["dose_mg"] == 0.5 and items[1]["daily_dose_mg"] == 1.0


def test_korean_frequencies_and_tablet_quantity_are_applied():
    text = "리스페달 2mg 2정 1일 2회, 아빌리파이 10mg 매일"
    items = app.test_client().post("/api/parse", json={"text": text}).get_json()["items"]
    assert items[0]["dose_mg"] == 4 and items[0]["daily_dose_mg"] == 8
    assert items[0]["frequency"] == "BID"
    assert items[1]["daily_dose_mg"] == 10 and items[1]["frequency"] == "QD"


def test_weekly_oral_schedule_is_not_averaged_into_a_daily_dose():
    item = app.test_client().post("/api/parse", json={"text": "risperidone 2mg weekly"}).get_json()["items"][0]
    assert item["ok"] is False and "주사제/주 단위" in item["error"]


def test_hourly_korean_and_regimen_frequencies_are_calculated_explicitly():
    text = "Risperdal 2mg q12h, Risperdal 2mg 2회/일, Risperdal 2mg 1-0-1, Risperdal 2mg q8h"
    items = app.test_client().post("/api/parse", json={"text": text}).get_json()["items"]
    assert [item["daily_dose_mg"] for item in items] == [4, 4, 4, 6]
    assert [item["frequency"] for item in items] == ["BID", "BID", "BID", "TID"]
    assert all(item["needs_review"] is False for item in items)


def test_daily_unit_and_quantity_before_strength_are_not_miscalculated():
    items = app.test_client().post(
        "/api/parse", json={"text": "Risperdal 4mg/day, Risperdal 2정 2mg BID"}
    ).get_json()["items"]
    assert items[0]["daily_dose_mg"] == 4 and items[0]["frequency"] == "QD"
    assert items[1]["dose_mg"] == 4 and items[1]["daily_dose_mg"] == 8


def test_patient_ids_stay_consistent_across_sheets_and_formula_text_is_safe():
    workbook = io.BytesIO()
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        pd.DataFrame({"patient_id": ["P01"], "medication": ["@Risperdal 2mg BID"]}).to_excel(writer, sheet_name="one", index=False)
        pd.DataFrame({"patient_id": ["P01", "P02"], "medication": ["Risperdal 2mg BID", "Abilify 10mg QD"]}).to_excel(writer, sheet_name="two", index=False)
    response = _upload(app.test_client(), workbook.getvalue(), "multi.xlsx")
    detailed = pd.read_excel(io.BytesIO(response.data), sheet_name="Detailed")
    assert detailed["patient"].tolist() == ["PATIENT_00001", "PATIENT_00001", "PATIENT_00002"]
    saved = load_workbook(io.BytesIO(response.data), data_only=False)
    original_column = [cell.value for cell in saved["Detailed"][1]].index("original") + 1
    originals = [saved["Detailed"].cell(row=row, column=original_column).value for row in range(2, saved["Detailed"].max_row + 1)]
    assert originals[0].startswith("'") and saved["Detailed"].cell(row=2, column=original_column).data_type != "f"
    assert "MethodInfo" in saved.sheetnames
    assert saved["Detailed"].freeze_panes == "A2"
    assert saved["Detailed"].auto_filter.ref == saved["Detailed"].dimensions
    assert saved["Detailed"].column_dimensions["E"].width > 11
    assert saved["Detailed"].row_dimensions[2].height >= 18


def test_drug_id_is_not_mistaken_for_patient_identifier():
    csv_bytes = b"drug_id,medication\nD001,Risperdal 2mg BID\n"
    response = _upload(app.test_client(), csv_bytes, "data.csv")
    detailed = pd.read_excel(io.BytesIO(response.data), sheet_name="Detailed")
    assert detailed.loc[0, "patient"] != "PATIENT_00001"


def test_every_registered_alias_parses_to_its_declared_standard_name():
    project = Path(__file__).resolve().parents[1]
    aliases = pd.read_csv(project / "lookup" / "drug_alias.csv")
    for row in aliases.itertuples():
        parsed = parse_medication(f"{row.alias} 1mg QD")
        assert parsed["drug"] == str(row.standard_name).casefold().strip()


def test_all_lookup_pairs_have_a_consistent_reciprocal():
    project = Path(__file__).resolve().parents[1]
    lookup = pd.read_csv(project / "lookup" / "master_lookup.csv")
    factors = {(row.method_id, row.source_drug, row.target_drug): row.factor for row in lookup.itertuples()}
    for row in lookup.itertuples():
        inverse = factors[(row.method_id, row.target_drug, row.source_drug)]
        assert abs(row.factor * inverse - 1) < 2e-6


def test_split_drug_dose_unit_and_frequency_columns_are_combined():
    csv_bytes = "환자번호,약물명,용량,단위,복용빈도\nP01,리스페달,2,mg,2\nP02,아빌리파이,10,mg,QD\n".encode("utf-8-sig")
    response = _upload(app.test_client(), csv_bytes, "분리열.csv")
    detailed = pd.read_excel(io.BytesIO(response.data), sheet_name="Detailed")
    assert detailed["drug"].tolist() == ["risperidone", "aripiprazole"]
    assert detailed["daily_dose_mg"].tolist() == [4, 10]
    assert detailed["source_row"].tolist() == [2, 3]
    assert detailed["medication_column"].tolist() == ["약물명", "약물명"]


def test_dose_unit_is_inferred_from_header_and_daily_dose_is_not_multiplied():
    csv_bytes = b"patient_id,drug,daily_dose_mg\nP01,risperidone,4\n"
    response = _upload(app.test_client(), csv_bytes, "daily.csv")
    detailed = pd.read_excel(io.BytesIO(response.data), sheet_name="Detailed")
    assert detailed.loc[0, "dose_mg"] == 4
    assert detailed.loc[0, "daily_dose_mg"] == 4


def test_daily_dose_column_is_not_multiplied_by_a_separate_frequency_column():
    csv_bytes = b"patient_id,drug,daily_dose_mg,frequency\nP01,risperidone,4,2\n"
    response = _upload(app.test_client(), csv_bytes, "daily_with_frequency.csv")
    detailed = pd.read_excel(io.BytesIO(response.data), sheet_name="Detailed")
    assert detailed.loc[0, "original"] == "risperidone 4mg QD"
    assert detailed.loc[0, "daily_dose_mg"] == 4


def test_patient_identifier_normalization_is_consistent_across_sheets():
    workbook = io.BytesIO()
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        pd.DataFrame({"patient_id": [1], "medication": ["Risperdal 1mg QD"]}).to_excel(writer, sheet_name="numeric", index=False)
        pd.DataFrame({"patient_id": ["1"], "medication": ["Abilify 10mg QD"]}).to_excel(writer, sheet_name="text", index=False)
    response = _upload(app.test_client(), workbook.getvalue(), "ids.xlsx")
    detailed = pd.read_excel(io.BytesIO(response.data), sheet_name="Detailed")
    assert detailed["patient"].tolist() == ["PATIENT_00001", "PATIENT_00001"]


def test_all_methods_are_exported_when_site_uses_all():
    csv_bytes = b"patient_id,medication\nP01,Risperdal 2mg BID\n"
    response = _upload(app.test_client(), csv_bytes, "all.csv", method="ALL")
    assert response.status_code == 200
    detailed = pd.read_excel(io.BytesIO(response.data), sheet_name="Detailed")
    assert {"CMD", "MED", "ED95", "DDD"}.issubset(set(detailed["method"]))
    assert detailed["equivalent_dose_mg"].notna().all()
    audit = pd.read_excel(io.BytesIO(response.data), sheet_name="AuditTrail")
    assert "CPZ_FGA" in audit.loc[0, "unavailable_methods"]


def test_common_korean_structured_column_names_are_detected():
    csv_bytes = "환자ID,제품명,1회용량,용량단위,일일횟수\nK01,리스페달,2,mg,2\n".encode("utf-8-sig")
    response = _upload(app.test_client(), csv_bytes, "병원추출.csv")
    assert response.status_code == 200
    detailed = pd.read_excel(io.BytesIO(response.data), sheet_name="Detailed")
    assert detailed.loc[0, "drug"] == "risperidone"
    assert detailed.loc[0, "daily_dose_mg"] == 4
    assert detailed.loc[0, "medication_column"] == "제품명"


def test_case_punctuation_and_curated_korean_typos_are_normalized():
    items = app.test_client().post(
        "/api/parse",
        json={"text": "RISP*ERDAL 2MG BID, 코티아핀 25mg QD"},
    ).get_json()["items"]
    assert items[0]["drug"] == "risperidone"
    assert items[1]["drug"] == "quetiapine"
    assert all("conversions" in item for item in items)


def test_short_ambiguous_korean_string_is_not_forced_to_a_drug():
    item = app.test_client().post("/api/parse", json={"text": "코틴 25mg QD"}).get_json()["items"][0]
    assert item["ok"] is False and item["needs_review"] is True


def test_sample_download_is_a_real_xlsx_file():
    response = app.test_client().get("/sample")
    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.data))
    assert workbook.active["B2"].value == "Risperdal"


def test_excel_with_exactly_five_sheets_is_accepted():
    workbook = io.BytesIO()
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        for index in range(1, 6):
            pd.DataFrame({
                "patient_id": [f"P{index:02d}"],
                "medication": ["Risperdal 1mg QD"],
            }).to_excel(writer, sheet_name=f"Sheet{index}", index=False)
    response = _upload(app.test_client(), workbook.getvalue(), "five_sheets.xlsx")
    assert response.status_code == 200
    detailed = pd.read_excel(io.BytesIO(response.data), sheet_name="Detailed")
    assert detailed["sheet"].nunique() == 5


def test_excel_with_more_than_five_sheets_is_rejected():
    workbook = io.BytesIO()
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        for index in range(1, 7):
            pd.DataFrame({"medication": ["Risperdal 1mg QD"]}).to_excel(
                writer, sheet_name=f"Sheet{index}", index=False
            )
    response = _upload(app.test_client(), workbook.getvalue(), "six_sheets.xlsx")
    assert response.status_code == 400
    assert "최대 5개" in response.get_data(as_text=True)
